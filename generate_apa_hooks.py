#!/usr/bin/env python3
"""
APA 2026 Hook Generator
========================
Usage:
  python3 generate_apa_hooks.py                    # All leads with notes in column L
  python3 generate_apa_hooks.py --row 5            # Single lead (hub row number)
  python3 generate_apa_hooks.py --auto-research    # Use Serper API to fill research notes
  python3 generate_apa_hooks.py --dry-run          # Preview without generating
  python3 generate_apa_hooks.py --sync             # Sync: remove from Hook Generator any leads deleted from Ben Leads Ready

ONE prompt per lead. Reads your notes, picks the APA SUBTHEME, writes the hook.
Email assembled from exact APA template — no AI on boilerplate.
"""

import os
import sys
import argparse
import subprocess
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

HUB_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'APA_2026_Combined_Hub.xlsx')
SERPER_API_KEY = os.environ.get('SERPER_API_KEY', '')

# ============================================================
# THE SINGLE PROMPT — picks subtheme, writes hook
# ============================================================
HOOK_PROMPT = """You are writing a hook paragraph for an outreach email about the APA 2026 Annual Meeting (American Psychiatric Association, May 16-20, 2026, San Francisco).

CRITICAL RULES:
- Use ONLY facts from the "Researcher Notes" below. Do NOT invent, assume, or hallucinate ANY information.
- If the notes mention multiple programs, pick the ONE that best connects to an APA 2026 subtheme.
- Name the actual program/initiative, be specific.
- NEVER use em dashes. No instances of the character sequence dash-dash or the long dash character anywhere in your output.

APA 2026 SUBTHEMES (you MUST pick the MOST SPECIFIC subtheme that fits):

1. Technology and AI
   - 'AI in Personalized Care': Wearable devices, mobile data, telehealth, virtual care, digital platforms integrated with neuroscience for personalized mental health treatments. USE THIS for any digital health, telehealth, virtual care, or health technology work.
   - 'AI Chatbots & Digital Companions': Risks and benefits of AI-driven emotional connections.
   - 'Ethical AI Design': Creating ethical, human-centered AI systems.

2. Workforce Empowerment and Mental Health Practice
   - 'Empowering The Psychiatric Workforce': "Taking Control of Our Practices One Step at a Time", the 2026 central theme.
   - 'Modernizing the Profession': New training, regulation, licensing pathways, campus professional development, workforce pipeline programs.
   - 'Combating Burnout': Improved, collaborative, and sustainable practice models.

3. Integrated Care and Systemic Changes
   - 'Whole-Person Treatment': ONLY use this when the work is specifically about integrating mental health into primary care or medical settings. Do NOT use as a catch-all.
   - 'Rethinking Child Support': ONLY use when the work is specifically focused on children, adolescents, or youth mental health programs.
   - 'Advocacy in Practice': Stepping into leadership roles to protect access to care and professional integrity.

4. Societal and Global Challenges
   - 'Environmental Impact': Mental health consequences of extreme weather and climate change.
   - 'Addressing Inequity and Polarization': Dismantling systemic racism, antisemitism, homophobia, transphobia; fostering inclusivity. Also: underserved communities, equity-focused care.
   - 'Global Perspectives': Cross-border collaboration on migration and human rights.

5. Emerging Clinical Trends
   - 'Psychedelic Treatment': Research and integration of psychedelic-assisted therapy.
   - 'Trauma-Informed Care': Any work focused on trauma survivors, PTSD, trauma-focused treatment, or care for trauma-affected populations. USE THIS whenever trauma is mentioned.
   - 'Neuroplasticity and Addiction': Any work on addiction, substance use disorder, recovery, or neuroplasticity research. USE THIS whenever addiction or substance use is mentioned.

SUBTHEME SELECTION RULES:
- DO NOT default to 'Whole-Person Treatment' unless the work is genuinely about integrating mental health into primary care.
- If the notes mention trauma or trauma survivors, use 'Trauma-Informed Care'.
- If the notes mention addiction, substance use, or recovery, use 'Neuroplasticity and Addiction'.
- If the notes mention virtual care, telehealth, or digital health, use 'AI in Personalized Care'.
- If the notes mention workforce training, campus programs, or professional development, use 'Modernizing the Profession'.
- If the notes mention underserved populations, equity, or diverse communities, use 'Addressing Inequity and Polarization'.

Institution: {institution}
Researcher Notes (YOUR ONLY SOURCE OF FACTS):
{notes}

Write a hook paragraph (3 sentences) following this EXACT structure:

SENTENCE 1: "We are keen to profile your work in [GENERAL AREA OF THEIR WORK, e.g. 'community-based mental health', 'integrated behavioral healthcare', 'addiction treatment and recovery'], and through our research we have been particularly impressed by [SPECIFIC PROGRAM/INITIATIVE NAME FROM NOTES]."
SENTENCE 2: One sentence describing what that specific program does, using ONLY facts from the notes, showing you know their work well.
SENTENCE 3: "Given that one of the main themes at this year's APA is '[SUBTHEME NAME]', we feel that highlighting [SPECIFIC ASPECT OF THEIR WORK THAT CONNECTS TO THIS SUBTHEME] would be particularly impactful, and we believe a film profiling your work would make for an incredibly valuable addition to the series."

CRITICAL RULES FOR TONE AND STYLE:
- Sentence 1 must NOT start with the institution name. Lead with "your work in".
- Sentence 1 starts BROAD (their general field/mission) then narrows to the specific program. This avoids pigeon-holing.
- When describing their work, use "the" not "a". You are DESCRIBING something they already know about, not EXPLAINING it to them. E.g. "the statewide effort" NOT "a statewide effort". "The 16-bed residential program" NOT "a 16-bed residential program".
- ALWAYS put the subtheme name in single apostrophes like 'Combating Burnout' or 'Trauma-Informed Care'.
- Do NOT repeat yourself. Each sentence must say something DIFFERENT. Sentence 1 introduces the general area and names the program. Sentence 2 describes what it does. Sentence 3 connects it to the APA theme.
- NEVER use em dashes anywhere in the output. Use commas, semicolons, or separate sentences instead.
- If no subtheme fits well, use the parent theme name instead.

After the paragraph, on a new line write:
SUBTHEME: [the specific subtheme name, e.g. "Trauma-Informed Care"]
PARENT: [the parent theme, e.g. "5. Emerging Clinical Trends"]

Output ONLY the hook paragraph, the SUBTHEME line, and the PARENT line."""

APA_EMAIL_TEMPLATE = """Dear Dr. {lastname},

I would like to schedule a call between you and Mark Rose, APA TV director, to discuss potentially highlighting {institution} in a pre-recorded video case study within the official broadcast at the American Psychiatric Association (APA) 2026 Annual Meeting in San Francisco (May 16-20, 2026) and online.

{hook}

As I'm sure you are aware, the APA Annual Meeting offers the largest audience of psychiatrists and mental health professionals at any meeting in the world. The APA has again partnered with WebsEdge, following 14 extremely successful years, to produce APA TV, the official broadcast of the APA Annual Meeting. This enables an important platform to showcase to the attendees some of the latest ground-breaking innovations across research, training, technology and patient care that are helping to shape the future of psychiatry and mental health.

As a key part of the APA TV broadcast, we will once again be highlighting to the attendees some companies, psychiatry departments and health systems that are doing great work within psychiatry and offer them a unique opportunity to profile their key research, recruitment initiatives and best practices in the form of a five-minute documentary feature.

Through our research, we are considering a number of companies, hospitals and health systems as potential candidates to sponsor these documentary features including {institution}, and I am keen to arrange a conversation between you and Mark Rose to make sure that there is a strong fit. I must emphasise that there is a cost involved in this opportunity to be profiled, which covers the production, distribution, and full ownership of the film and all additional footage.

As part of this project, any partner profiled in this way will retain the rights to the finished film and all footage we shoot. Mark will of course run through this and all logistics with you in more detail during the call.

In advance of the conversation, it would be useful for you to have a look at one or two of the groups that we profiled at recent APA Annual Meetings as this will give you an idea on the style of film we would produce with you. You can see a few of those films here:

{film_links}

As such, please could you email back with some suitable times over the next few days when Mark can call you to discuss this? He will be in meetings for the majority of today, but is fairly open over the next few days if you can suggest a couple of times for an initial call?

I look forward to hearing back from you with a convenient time to speak.

Best wishes,

Ben"""


def serper_search(query):
    import requests
    headers = {"X-API-KEY": SERPER_API_KEY, "Content-Type": "application/json"}
    try:
        resp = requests.post("https://google.serper.dev/search",
                           json={"q": query, "num": 5}, headers=headers, timeout=15)
        resp.raise_for_status()
        return '\n'.join(f"- {item.get('title', '')}: {item.get('snippet', '')}"
                        for item in resp.json().get('organic', [])[:5])
    except Exception as e:
        return f"Search failed: {e}"


def serper_research_lead(institution, area):
    queries = [
        f'"{institution}" mental health programs initiatives 2024 2025',
        f'"{institution}" {area} psychiatry new',
    ]
    results = []
    for q in queries:
        r = serper_search(q)
        if r and 'failed' not in r.lower():
            results.append(f"Query: {q}\n{r}")
    return '\n\n'.join(results) if results else "No results found"


def generate_hook_text(prompt):
    try:
        result = subprocess.run(
            ['claude', '-p', prompt, '--allowedTools', ''],
            capture_output=True, text=True, timeout=120
        )
        if result.returncode == 0:
            return result.stdout.strip()
        return f"ERROR: {result.stderr[:200]}"
    except FileNotFoundError:
        return "ERROR: CLI tool not found"
    except subprocess.TimeoutExpired:
        return "ERROR: Generation timed out"
    except Exception as e:
        return f"ERROR: {e}"


def process_lead(ws, row, auto_research=False):
    name = ws.cell(row, 2).value
    institution = ws.cell(row, 4).value or ''
    area = ws.cell(row, 6).value or ''
    notes = ws.cell(row, 12).value or ''  # Column L = researcher notes
    film1 = ws.cell(row, 17).value or ''  # Column Q = Film 1
    film2 = ws.cell(row, 18).value or ''  # Column R = Film 2
    film3 = ws.cell(row, 19).value or ''  # Column S = Film 3

    if not name:
        return False, "No name found"

    if auto_research and not notes.strip():
        print(f"  Researching {institution}...")
        if not SERPER_API_KEY:
            return False, "SERPER_API_KEY not set (export SERPER_API_KEY=your_key)"
        research = serper_research_lead(institution, area)
        ws.cell(row, 12, research)
        ws.cell(row, 12).fill = PatternFill('solid', fgColor='FFF2CC')
        notes = research

    if not notes.strip():
        return False, "No researcher notes, paste notes in column L first"

    # ONE PROMPT: reads notes → picks subtheme → writes hook
    print(f"  Generating hook for {name} at {institution}...")
    prompt = HOOK_PROMPT.format(institution=institution, notes=notes)
    result = generate_hook_text(prompt)
    if result.startswith("ERROR"):
        return False, result

    # Parse output
    lines = result.strip().split('\n')
    subtheme = ''
    parent_theme = ''
    hook_lines = []
    for line in lines:
        stripped = line.strip()
        if stripped.upper().startswith('SUBTHEME:'):
            subtheme = stripped.split(':', 1)[1].strip()
        elif stripped.upper().startswith('PARENT:'):
            parent_theme = stripped.split(':', 1)[1].strip()
        elif stripped:
            hook_lines.append(stripped)

    hook_paragraph = ' '.join(hook_lines) if hook_lines else result.split('\n')[0]

    # Strip any em dashes from hook
    hook_paragraph = hook_paragraph.replace('\u2014', ',').replace('\u2013', ',')

    ws.cell(row, 14, subtheme)       # APA Subtheme (generated)
    ws.cell(row, 15, parent_theme)   # APA Parent Theme
    ws.cell(row, 16, hook_paragraph) # Generated Hook

    # ASSEMBLE EMAIL, pure template, no AI
    lastname = name.split()[-1] if name else 'Name'
    film_links = '\n'.join(f.strip() for f in [film1, film2, film3] if f and f.strip())

    full_email = APA_EMAIL_TEMPLATE.format(
        lastname=lastname, institution=institution,
        hook=hook_paragraph, film_links=film_links
    )

    ws.cell(row, 20, full_email)  # Full Draft Email
    ws.cell(row, 21, 'Done')
    ws.cell(row, 21).fill = PatternFill('solid', fgColor='C6EFCE')
    return True, "Complete"


def sync_tabs(wb):
    """Remove leads from Hook Generator that no longer exist in Ben Leads Ready."""
    leads_ws = wb['Ben Leads Ready']
    hook_ws = wb['Hook Generator']

    # Build set of (name, institution) from Ben Leads Ready
    # Contact Name = col 8, Institution = col 4 in Ben Leads Ready
    active_leads = set()
    for r in range(2, leads_ws.max_row + 1):
        name = (leads_ws.cell(r, 8).value or '').strip().lower()
        inst = (leads_ws.cell(r, 4).value or '').strip().lower()
        if name:
            active_leads.add((name, inst))

    # Check Hook Generator and mark rows to delete
    # Contact Name = col 2, Institution = col 4 in Hook Generator
    rows_to_delete = []
    for r in range(2, hook_ws.max_row + 1):
        name = (hook_ws.cell(r, 2).value or '').strip().lower()
        inst = (hook_ws.cell(r, 4).value or '').strip().lower()
        if name and (name, inst) not in active_leads:
            rows_to_delete.append(r)

    # Delete from bottom up to preserve row numbers
    removed = 0
    for r in reversed(rows_to_delete):
        lead_name = hook_ws.cell(r, 2).value
        lead_inst = hook_ws.cell(r, 4).value
        print(f"  Removing row {r}: {lead_name} — {lead_inst}")
        hook_ws.delete_rows(r)
        removed += 1

    # Re-number the Row# column
    for r in range(2, hook_ws.max_row + 1):
        hook_ws.cell(r, 1, r - 1)

    print(f"\n{'='*60}")
    print(f"Sync complete: {removed} leads removed from Hook Generator")
    print(f"Hook Generator now has {hook_ws.max_row - 1} leads (matching Ben Leads Ready: {len(active_leads)})")
    print(f"{'='*60}")


def main():
    parser = argparse.ArgumentParser(description='APA 2026 Hook Generator')
    parser.add_argument('--row', type=int, help='Hub row number to process')
    parser.add_argument('--auto-research', action='store_true', help='Use Serper for research')
    parser.add_argument('--dry-run', action='store_true', help='Preview without generating')
    parser.add_argument('--sync', action='store_true', help='Sync Hook Generator with Ben Leads Ready (removes deleted leads)')
    args = parser.parse_args()

    if not os.path.exists(HUB_FILE):
        print(f"Error: Hub file not found at {HUB_FILE}")
        sys.exit(1)

    wb = load_workbook(HUB_FILE)
    ws = wb['Hook Generator']

    if args.sync:
        sync_tabs(wb)
        wb.save(HUB_FILE)
        return

    if args.row:
        rows = [args.row]
    else:
        rows = []
        for r in range(2, ws.max_row + 1):
            status = ws.cell(r, 21).value or 'Pending'
            notes = ws.cell(r, 12).value or ''
            name = ws.cell(r, 2).value
            if name and status in ('Pending', '', None):
                if notes.strip() or args.auto_research:
                    rows.append(r)

    if not rows:
        print("No leads to process.")
        print("Paste researcher notes into column L of the Hook Generator tab in APA_2026_Combined_Hub.xlsx")
        return

    print(f"\n{'='*60}")
    print(f"APA 2026 Hook Generator — {len(rows)} leads")
    print(f"{'='*60}\n")

    if args.dry_run:
        for r in rows:
            name = ws.cell(r, 2).value
            inst = ws.cell(r, 4).value
            notes = ws.cell(r, 12).value or ''
            print(f"  Row {r}: {name} - {inst} {'has notes' if notes.strip() else 'NO notes'}")
        return

    ok_count, fail_count = 0, 0
    for i, r in enumerate(rows):
        name = ws.cell(r, 2).value
        inst = ws.cell(r, 4).value
        print(f"[{i+1}/{len(rows)}] {name} — {inst}")
        ok, msg = process_lead(ws, r, auto_research=args.auto_research)
        print(f"  {'done' if ok else 'SKIP'}: {msg}")
        if ok: ok_count += 1
        else: fail_count += 1
        wb.save(HUB_FILE)

    print(f"\n{'='*60}")
    print(f"Done: {ok_count} generated, {fail_count} skipped")
    print(f"Saved: {HUB_FILE}")
    print(f"{'='*60}")


if __name__ == '__main__':
    main()
